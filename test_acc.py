import unittest
from unittest.mock import patch, MagicMock
import account
from db.db import get_db_connection
import tkinter
import openpyxl
from openpyxl.utils import get_column_letter


class TestAccountFunctions(unittest.TestCase):
    def test_hash_password(self):
        password = "secret"
        hashed = account.hash_password(password)
        self.assertNotEqual(password, hashed)
        self.assertEqual(len(hashed), 64)

    def test_verify_password_success(self):
        password = "password123"
        hashed = account.hash_password(password)
        self.assertTrue(account.verify_password(password, hashed))

    def test_verify_password_fail(self):
        password = "password123"
        wrong_password = "wrongpassword"
        hashed = account.hash_password(password)
        self.assertFalse(account.verify_password(wrong_password, hashed))

    def test_generate_reset_code_length(self):
        reset_code = account.generate_reset_code()
        self.assertEqual(len(reset_code), 12)


class TestAccountGUI(unittest.TestCase):
    test_results = []  # Class-level list to store test results

    @classmethod
    def setUpClass(cls):
        cls.test_results = []

    def setUp(self):
        self.test_name = self.id().split('.')[-1]  # Get the name of the current test

    def tearDown(self):
        # Record the outcome of each test
        outcome = self._outcome.errors
        if outcome:
            result = 'Fail'
            for error in outcome:
                if error[0] is not None:
                    result += f" ({error[0].__name__}: {error[1]})"
        else:
            result = 'Pass'
        TestAccountGUI.test_results.append({'Test Case': self.test_name, 'Result': result})

    @classmethod
    def tearDownClass(cls):
        # Write the test results to an Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Account Test Results"

        # Write headers
        headers = ['Test Case', 'Result']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

        # Write test results
        for row_num, result in enumerate(cls.test_results, 2):
            sheet[f"A{row_num}"] = result['Test Case']
            sheet[f"B{row_num}"] = result['Result']

        workbook.save("account_test_results.xlsx")
        print("\nTest results written to account_test_results.xlsx")

    
    
    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_login_fail_wrong_password(self, mock_showerror, mock_get_db_connection):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = {'id': 1, 'email': 'test@gmail.com',
                                            'password_hash': account.hash_password("CorrectPassword")}  # Mật khẩu đúng
        mock_conn.cursor.return_value = mock_cursor
        mock_get_db_connection.return_value = mock_conn

        result = account.login("test@gmail.com", "1")  # Nhập sai mật khẩu
        self.assertIsNone(result)
        mock_showerror.assert_called_once_with("Lỗi", "Sai email hoặc mật khẩu.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_login_fail_wrong_email(self, mock_showerror, mock_get_db_connection):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = None
        mock_conn.cursor.return_value = mock_cursor
        mock_get_db_connection.return_value = mock_conn

        result = account.login("asd@gmail.com", "Test123@")  # Nhập sai email
        self.assertIsNone(result)
        mock_showerror.assert_called_once_with("Lỗi", "Sai email hoặc mật khẩu.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_login_fail_empty_password(self, mock_showerror, mock_get_db_connection):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = {'id': 1, 'email': 'test@gmail.com',
                                            'password_hash': account.hash_password("Test123@")}
        mock_conn.cursor.return_value = mock_cursor
        mock_get_db_connection.return_value = mock_conn

        result = account.login("test@gmail.com", "")  # Bỏ trống mật khẩu
        self.assertIsNone(result)
        mock_showerror.assert_called_once_with("Lỗi", "Sai email hoặc mật khẩu.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_login_fail_empty_email(self, mock_showerror, mock_get_db_connection):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = None
        mock_conn.cursor.return_value = mock_cursor
        mock_get_db_connection.return_value = mock_conn

        result = account.login("", "1")  # Bỏ trống email
        self.assertIsNone(result)
        mock_showerror.assert_called_once_with("Lỗi", "Sai email hoặc mật khẩu.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_login_fail_empty_email_and_password(self, mock_showerror, mock_get_db_connection):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = None
        mock_conn.cursor.return_value = mock_cursor
        mock_get_db_connection.return_value = mock_conn

        result = account.login("", "")  # Bỏ trống email và mật khẩu
        self.assertIsNone(result)
        mock_showerror.assert_called_once_with("Lỗi", "Sai email hoặc mật khẩu.")

    @patch('account.get_db_connection')
    def test_login_success(self, mock_get_db_connection):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = {'id': 1, 'email': 'test@gmail.com',
                                            'password_hash': account.hash_password("Test123@")}
        mock_conn.cursor.return_value = mock_cursor
        mock_get_db_connection.return_value = mock_conn

        result = account.login("test@gmail.com", "Test123@")  # Nhập đúng email và mật khẩu
        self.assertEqual(result, "test@gmail.com")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_register_user_exists(self, mock_showerror, mock_get_db_connection):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = 1
        mock_conn.cursor.return_value = mock_cursor
        mock_get_db_connection.return_value = mock_conn

        result = account.register("asd@gmail.com", "password", "password")
        self.assertFalse(result)
        mock_showerror.assert_called_once_with("Lỗi", "Email 'asd@gmail.com' đã tồn tại.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_register_fail_invalid_email(self, mock_showerror, mock_get_db_connection):
        result = account.register("asd@asd.com", "1", "1")
        self.assertFalse(result)
        mock_showerror.assert_called_once_with("Lỗi", "Định dạng email không hợp lệ.\nVui lòng sử dụng @gmail.com hoặc @sgu.edu.vn.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_register_fail_password_too_short(self, mock_showerror, mock_get_db_connection):
        result = account.register("test1@gmail.com", "1", "1")
        self.assertFalse(result)
        mock_showerror.assert_called_once_with("Lỗi", "Mật khẩu không hợp lệ.\nMật khẩu phải có ít nhất 8 ký tự, bao gồm chữ hoa, chữ thường, số và ký tự đặc biệt.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_register_success(self, mock_showerror, mock_get_db_connection):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = 0
        mock_conn.cursor.return_value = mock_cursor
        mock_get_db_connection.return_value = mock_conn

        result = account.register("test1@gmail.com", "Test123@", "Test123@")
        self.assertTrue(result)

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_register_password_mismatch(self, mock_showerror, mock_get_db_connection):
        result = account.register("email@example.com", "password", "wrongpassword")
        self.assertFalse(result)
        mock_showerror.assert_called_once_with("Lỗi", "Mật khẩu không khớp.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_register_invalid_email_format(self, mock_showerror, mock_get_db_connection):
        result = account.register("invalid_email", "Password123!", "Password123!")
        self.assertFalse(result)
        mock_showerror.assert_called_once_with("Lỗi", "Định dạng email không hợp lệ.\nVui lòng sử dụng @gmail.com hoặc @sgu.edu.vn.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_register_invalid_password_format(self, mock_showerror, mock_get_db_connection):
        result = account.register("email@gmail.com", "weak", "weak")
        self.assertFalse(result)
        mock_showerror.assert_called_once_with("Lỗi", "Mật khẩu không hợp lệ.\nMật khẩu phải có ít nhất 8 ký tự, bao gồm chữ hoa, chữ thường, số và ký tự đặc biệt.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_forgot_password_user_not_found(self, mock_showerror, mock_get_db_connection):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = None
        mock_conn.cursor.return_value = mock_cursor
        mock_get_db_connection.return_value = mock_conn

        result = account.forgot_password("nonexistent@gmail.com", "Test123@")
        self.assertFalse(result)
        mock_showerror.assert_called_once_with("Lỗi", " ")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_forgot_password_invalid_email(self, mock_showerror, mock_get_db_connection):
        result = account.forgot_password("asd@asd.com", "Test123!")
        self.assertFalse(result)
        mock_showerror.assert_called_once_with("Lỗi", "Email không hợp lệ.")

    @patch('account.get_db_connection')
    @patch('tkinter.messagebox.showerror')
    def test_forgot_password_invalid_password(self, mock_showerror, mock_get_db_connection):
        result = account.forgot_password("test@sgu.edu.vn", "1")
        self.assertFalse(result)
        mock_showerror.assert_called_once_with("Lỗi", "Mật khẩu phải ít nhất 8 ký tự, có chữ hoa, chữ thường, số và ký tự đặc biệt.")

    @patch('account.get_db_connection')
    def test_forgot_password_success(self, mock_get_db_connection):
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_cursor.fetchone.return_value = {'id': 1, 'email': 'test@sgu.edu.vn',
                                            'password_hash': 'old_hash'}
        mock_conn.cursor.return_value = mock_cursor
        mock_get_db_connection.return_value = mock_conn

        result = account.forgot_password("test@sgu.edu.vn", "Test123!")
        self.assertTrue(result)


if __name__ == '__main__':
    unittest.main()